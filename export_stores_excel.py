"""יצוא חנויות לאקסל — מצפון לדרום, בתוך כל עיר סדר נסיעה נוח + היסטוריית ביקורים"""
import csv, sys, math, io, requests
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from visit_tracker import get_all_visit_stats, urgency_color_hex
sys.stdout.reconfigure(encoding="utf-8")

# ── טען נתונים מ-GitHub ────────────────────────────────────────────────────────
BASE = "https://raw.githubusercontent.com/yonatan-avshalomov/-whatsapp-bot/main"

def fetch_csv(url):
    r = requests.get(url, timeout=15)
    return list(csv.DictReader(io.StringIO(r.content.decode("utf-8-sig"))))

try:
    stores     = fetch_csv(f"{BASE}/stores.csv")
    deliveries = fetch_csv(f"{BASE}/senzey_data.csv")
    manual_v   = fetch_csv(f"{BASE}/manual_visits.csv")
    visit_stats = get_all_visit_stats(stores, deliveries, manual_v)
    print(f"✅ היסטוריית ביקורים: {sum(1 for d in visit_stats.values() if d['days_since'] is not None)} חנויות בוקרו")
except Exception as e:
    print(f"⚠️  לא ניתן לטעון היסטוריית ביקורים: {e}")
    # fallback — טען מקובץ מקומי
    with open("stores.csv", encoding="utf-8-sig") as f:
        stores = list(csv.DictReader(f))
    visit_stats = {}

def safe_float(v, default=0.0):
    try:
        return float(v) if v and str(v).strip() not in ("", "0", "0.0", "None") else default
    except:
        return default

# ── Nearest-Neighbor בתוך עיר ───────────────────────────────────────────────
def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1))*math.cos(math.radians(lat2))*math.sin(dlon/2)**2
    return R * 2 * math.asin(math.sqrt(a))

def nearest_neighbor_route(city_stores):
    """מסלול nearest-neighbor — מתחיל מהצפוני ביותר, עובר לקרוב הבא"""
    if len(city_stores) <= 2:
        return sorted(city_stores, key=lambda s: -safe_float(s.get("lat")))
    remaining = list(city_stores)
    # התחל מהצפוני ביותר (lat גבוה = צפון)
    current = max(remaining, key=lambda s: safe_float(s.get("lat")))
    route = [current]
    remaining.remove(current)
    while remaining:
        clat = safe_float(current.get("lat"))
        clon = safe_float(current.get("lon"))
        nxt = min(remaining,
                  key=lambda s: haversine(clat, clon,
                                          safe_float(s.get("lat")),
                                          safe_float(s.get("lon"))))
        route.append(nxt)
        remaining.remove(nxt)
        current = nxt
    return route

# ── סנן כפילויות מספר הזמנה (אותו סניף, מספר שונה בכל תעודה) ─────────────────
def is_order_duplicate(name):
    """מחזיר True אם זה רק מספר הזמנה — לא סניף עצמאי"""
    markers = [":מספר הזמנה", "מספר הזמנה", " מספר", "- הזמנה מספר", "הזמנה מספר"]
    n = name.strip()
    return any(m in n for m in markers) or n.endswith(" מספר")

stores = [s for s in stores if not is_order_duplicate(s.get("name",""))]

# ── קבץ לפי עיר ──────────────────────────────────────────────────────────────
by_city = {}
for s in stores:
    city = s.get("city","").strip() or "לא ידוע"
    by_city.setdefault(city, []).append(s)

# ממוצע lat לכל עיר → מיון צפון→דרום
def city_avg_lat(city):
    lats = [safe_float(s.get("lat")) for s in by_city[city]
            if safe_float(s.get("lat")) > 0]
    return sum(lats)/len(lats) if lats else 0

cities_sorted = sorted(by_city.keys(), key=city_avg_lat, reverse=True)  # צפון → דרום

# ── צבעי רשתות ───────────────────────────────────────────────────────────────
CHAIN_COLORS = {
    "שילב":   {"header": "1F4E79", "row_even": "D6E4F0", "row_odd": "EBF4FB"},
    "מכבי":   {"header": "375623", "row_even": "D9EAD3", "row_odd": "EBF5E6"},
    "ניצת":   {"header": "7B3F00", "row_even": "FCE4D6", "row_odd": "FEF0E7"},
    "other":  {"header": "404040", "row_even": "EFEFEF", "row_odd": "F8F8F8"},
}

def chain_colors(chain_name):
    for key in CHAIN_COLORS:
        if key in (chain_name or ""):
            return CHAIN_COLORS[key]
    return CHAIN_COLORS["other"]

# ── בנה Excel ─────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "חנויות מצפון לדרום"
ws.sheet_view.rightToLeft = True

# כותרות
HEADERS = ["#", "עיר", "שם חנות", "רשת", "כתובת", "טלפון", "ביקור אחרון", "ימים", "ק\"מ מהוד השרון"]
col_widths = [5, 16, 36, 10, 28, 14, 13, 7, 12]

HOME_LAT, HOME_LON = 32.150, 34.893  # הוד השרון

# כותרת ראשית
ws.merge_cells("A1:I1")
title_cell = ws["A1"]
title_cell.value = "רשימת חנויות — מצפון לדרום  |  סדר נסיעה נוח בתוך כל עיר"
title_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor="1A1A2E")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# כותרות עמודות
for col, (h, w) in enumerate(zip(HEADERS, col_widths), 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2C3E50")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[2].height = 20

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row_num = 3
store_counter = 0

for city in cities_sorted:
    city_stores = by_city[city]
    route = nearest_neighbor_route(city_stores)

    # כותרת עיר
    city_lat = city_avg_lat(city)
    ws.merge_cells(f"A{row_num}:I{row_num}")
    city_cell = ws.cell(row=row_num, column=1,
                        value=f"📍  {city}  ({len(route)} חנויות)  |  {city_lat:.2f}°N")
    city_cell.font = Font(name="Arial", bold=True, size=11, color="1A1A2E")
    city_cell.fill = PatternFill("solid", fgColor="BDC3C7")
    city_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[row_num].height = 18
    row_num += 1

    for i, s in enumerate(route, 1):
        store_counter += 1
        name  = s.get("name", "")
        chain = s.get("chain", "")
        colors = chain_colors(chain)

        # ── נתוני ביקור ──
        vstats      = visit_stats.get(name, {})
        days_since  = vstats.get("days_since")
        last_date   = vstats.get("last_date_str", "—")

        # ── צבע שורה: מבוסס על דחיפות אם יש, אחרת לפי רשת ──
        if days_since is not None:
            visit_hex  = urgency_color_hex(days_since)
            fill_color = visit_hex
        else:
            fill_color = colors["row_even"] if i % 2 == 0 else colors["row_odd"]

        dist = haversine(HOME_LAT, HOME_LON,
                         safe_float(s.get("lat"), HOME_LAT),
                         safe_float(s.get("lon"), HOME_LON))

        vals = [
            store_counter,
            city,
            name,
            chain,
            s.get("address", ""),
            s.get("phone", ""),
            last_date,
            days_since if days_since is not None else "—",
            f"{dist:.1f}",
        ]
        aligns = ["center","right","right","center","right","center","center","center","center"]

        for col, (val, align) in enumerate(zip(vals, aligns), 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = border
        ws.row_dimensions[row_num].height = 15
        row_num += 1

# הקפא שורות כותרת
ws.freeze_panes = "A3"

# פילטר אוטומטי
ws.auto_filter.ref = f"A2:I{row_num-1}"

# גיליון סיכום לפי רשת
ws2 = wb.create_sheet("סיכום לפי רשת")
ws2.sheet_view.rightToLeft = True
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 20

summary_header = ["רשת", "מספר חנויות", "ערים"]
for col, h in enumerate(summary_header, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2C3E50")
    c.alignment = Alignment(horizontal="center")

chains_summary = {}
for s in stores:
    ch = s.get("chain","לא ידוע") or "לא ידוע"
    city = s.get("city","") or ""
    chains_summary.setdefault(ch, {"count": 0, "cities": set()})
    chains_summary[ch]["count"] += 1
    if city:
        chains_summary[ch]["cities"].add(city)

for r, (ch, data) in enumerate(sorted(chains_summary.items(),
                                       key=lambda x: -x[1]["count"]), 2):
    ws2.cell(row=r, column=1, value=ch)
    ws2.cell(row=r, column=2, value=data["count"])
    ws2.cell(row=r, column=3, value=f"{len(data['cities'])} ערים")

# שמור
out = "stores_north_to_south.xlsx"
wb.save(out)
print(f"נשמר: {out}")
print(f"סה\"כ: {store_counter} חנויות | {len(cities_sorted)} ערים")
print("ערים מצפון לדרום:")
for c in cities_sorted[:10]:
    lat = city_avg_lat(c)
    print(f"  {lat:.2f}°  {c}  ({len(by_city[c])} חנויות)")
print("  ...")
for c in cities_sorted[-5:]:
    lat = city_avg_lat(c)
    print(f"  {lat:.2f}°  {c}  ({len(by_city[c])} חנויות)")
